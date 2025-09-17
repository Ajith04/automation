import re
import zipfile
import xml.etree.ElementTree as ET
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter
import random

# ---------- CONFIG ----------
TARGET_SHEETS = ["AKUN", "WAMA", "GALAXEA"]
TARGET_RED = "FFC00000"
HEADERS = ["Event", "Resource", "Configuration", "Date", "Start Time", "End Time"]
HEADER_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
HEADER_FONT = Font(bold=True)
YEAR_FOR_OUTPUT = 2025  # you can set datetime.now().year if you want dynamic

# XML namespace used in sheet XML
NS = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
RELS_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


# ---------- HELPERS ----------
def safe_str(v):
    return "" if v is None else str(v).strip()


def get_rgb(cell):
    if cell is None:
        return None
    fill = getattr(cell, "fill", None)
    if not fill:
        return None
    color = getattr(fill, "start_color", None)
    if color is None:
        return None
    if hasattr(color, "rgb") and color.rgb:
        return str(color.rgb).upper()
    return None


def is_red(cell):
    return get_rgb(cell) == TARGET_RED


def add_headers(ws):
    for col_idx, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=col_idx, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL


def clean_instructor_name(name):
    if not name:
        return None
    return re.sub(r"\s*\(.*?\)\s*", "", str(name)).strip()


def get_light_fill():
    colors = [
        "FFFFE5CC", "FFE5FFCC", "FFCCFFE5", "FFCCE5FF",
        "FFFFCCFF", "FFE5CCFF", "FFFFCCCC", "FFCCFFFF"
    ]
    hexcolor = random.choice(colors)
    return PatternFill(start_color=hexcolor, end_color=hexcolor, fill_type="solid")


# ---------- XML Parsing (dataValidations) ----------
def _get_sheet_file_map(xlsx_file):
    """
    Return dict: sheet_name -> sheet_xml_path (e.g. 'xl/worksheets/sheet2.xml')
    """
    with zipfile.ZipFile(xlsx_file, "r") as z:
        # parse workbook.xml to get sheet name -> r:id
        wb_xml = z.read("xl/workbook.xml")
        root = ET.fromstring(wb_xml)
        sheets_el = root.find("main:sheets", NS)
        sheet_to_rid = {}
        if sheets_el is not None:
            rid_attr = "{%s}id" % RELS_NS
            for s in sheets_el.findall("main:sheet", NS):
                name = s.get("name")
                rid = s.get(rid_attr)
                if name and rid:
                    sheet_to_rid[name] = rid

        # parse relationships to map rId -> Target path
        rels_xml = z.read("xl/_rels/workbook.xml.rels")
        rroot = ET.fromstring(rels_xml)
        rid_to_target = {}
        for rel in rroot.findall("Relationship"):
            rid = rel.get("Id") or rel.get("Id".lower())
            target = rel.get("Target")
            if rid and target:
                # canonicalize path (target is relative to xl/)
                rid_to_target[rid] = "xl/" + target.lstrip("/")

        # build sheet_name -> sheet_file mapping
        sheet_map = {}
        for name, rid in sheet_to_rid.items():
            target = rid_to_target.get(rid)
            if target:
                sheet_map[name] = target
        return sheet_map


def _expand_sqref_to_cells(sqref):
    """
    Given a sqref string (e.g. 'E2:E10' or 'E2 E4' or 'E2'), return list of coords ['E2','E3',...]
    """
    coords = []
    parts = re.split(r"\s+", sqref.strip())
    for part in parts:
        if ":" in part:
            start, end = part.split(":")
            start_col, start_row = coordinate_from_string(start)
            end_col, end_row = coordinate_from_string(end)
            start_row = int(start_row)
            end_row = int(end_row)
            start_col_idx = column_index_from_string(start_col)
            end_col_idx = column_index_from_string(end_col)
            for cidx in range(start_col_idx, end_col_idx + 1):
                for r in range(start_row, end_row + 1):
                    coords.append(f"{get_column_letter(cidx)}{r}")
        else:
            coords.append(part)
    return coords


def parse_data_validations_xlsx(xlsx_file):
    """
    Parse workbook xmls and return dict:
      { sheet_name: { 'A2': formula1_text, 'B3': formula1_text, ... }, ... }
    """
    sheet_file_map = _get_sheet_file_map(xlsx_file)
    results = {name: {} for name in sheet_file_map.keys()}

    with zipfile.ZipFile(xlsx_file, "r") as z:
        for sheet_name, sheet_file in sheet_file_map.items():
            if sheet_file not in z.namelist():
                continue
            xml_content = z.read(sheet_file)
            root = ET.fromstring(xml_content)
            dvs = root.findall(".//main:dataValidation", NS)
            for dv in dvs:
                sqref = dv.get("sqref")
                formula_el = dv.find("main:formula1", NS)
                formula_text = formula_el.text if formula_el is not None else None
                if not sqref or not formula_text:
                    continue
                cell_coords = _expand_sqref_to_cells(sqref)
                for coord in cell_coords:
                    # keep first formula if duplicate; but you could override if needed
                    if coord not in results[sheet_name]:
                        results[sheet_name][coord] = formula_text
    return results


# ---------- Dropdown resolution helpers ----------
def resolve_formula_to_values(wb, formula):
    """
    Given a formula string (e.g. '"08:00 - 10:00,17:00 - 19:00"' or '=Availability!B2:B20'),
    return list of option strings.
    wb is openpyxl workbook instance (data_only=True recommended).
    """
    if not formula:
        return []

    formula = formula.strip()
    # inline list
    if formula.startswith('"') and formula.endswith('"'):
        inner = formula.strip('"')
        items = [x.strip() for x in inner.split(",") if x.strip()]
        return items

    # remove leading '=' if present
    if formula.startswith("="):
        ref = formula[1:].strip()
    else:
        ref = formula

    # handle cross-sheet ref like 'Availability!B2:B20' (sheet name might be quoted)
    if "!" in ref:
        sheet_part, rng = ref.split("!", 1)
        sheet_part = sheet_part.strip().strip("'").strip('"')
        rng = rng.replace("$", "")
        if sheet_part in wb.sheetnames:
            ws = wb[sheet_part]
            try:
                cells = ws[rng]
            except Exception:
                # fallback: return empty
                return []
            values = []
            # ws[rng] returns tuple(s)
            for row in cells:
                for c in row:
                    if c.value is not None:
                        values.append(str(c.value).strip())
            return list(dict.fromkeys(values))

    # if ref looks like same-sheet range (e.g. 'B2:B20')
    if ":" in ref:
        # we'll try to find it on all sheets - but normally this branch is for same-sheet ranges;
        # caller should pass appropriate sheet context if needed. As a safe fallback attempt
        values = []
        for ws in wb.worksheets:
            try:
                cells = ws[ref]
            except Exception:
                continue
            for row in cells:
                for c in row:
                    if c.value is not None:
                        values.append(str(c.value).strip())
        return list(dict.fromkeys(values))

    # named range?
    if ref in wb.defined_names:
        values = []
        try:
            dests = wb.defined_names[ref].destinations
            for title, area in dests:
                if title in wb.sheetnames:
                    ws = wb[title]
                    for row in ws[area]:
                        for c in row:
                            if c.value is not None:
                                values.append(str(c.value).strip())
            return list(dict.fromkeys(values))
        except Exception:
            return []

    return []


# ---------- STAFF LOADER ----------
def preload_staff(staff_file):
    wb = load_workbook(staff_file, data_only=True)
    result = {}
    for sheet in TARGET_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        headers = [safe_str(c.value).lower() for c in ws[1]]
        try:
            pri_idx = headers.index("priority") + 1
        except ValueError:
            pri_idx = 1
        instr_start_col = pri_idx + 1
        sheet_map = {}
        for col in range(instr_start_col, ws.max_column + 1):
            instr_name = clean_instructor_name(safe_str(ws.cell(1, col).value))
            if not instr_name:
                continue
            r = 2
            while r <= ws.max_row:
                val_cell = ws.cell(r, col)
                val = safe_str(val_cell.value)
                if not val:
                    break
                if is_red(val_cell):
                    r += 1
                    continue
                sheet_map.setdefault(val, []).append(instr_name)
                r += 1
        result[sheet] = sheet_map
    return result


# ---------- MAIN ----------
def generate_output(events_file, staff_file, output_file):
    # preload instructors
    instructors_map = preload_staff(staff_file)

    # parse data validations from XML (gives mapping sheet -> { coord: formula })
    dv_map = parse_data_validations_xlsx(events_file)

    # load workbook to read actual cell values (data_only to read evaluated values)
    wb_src = load_workbook(events_file, data_only=True)
    wb_out = Workbook()
    wb_out.remove(wb_out.active)
    event_color_cache = {}

    for sheet_name in wb_src.sheetnames:
        if sheet_name.upper() not in TARGET_SHEETS:
            continue

        ws_src = wb_src[sheet_name]
        ws_out = wb_out.create_sheet(sheet_name)
        add_headers(ws_out)

        # map headers in the source sheet
        header_map = {}
        for c in range(1, ws_src.max_column + 1):
            hv = safe_str(ws_src.cell(1, c).value).lower()
            if hv:
                header_map[hv] = c

        resort_col = header_map.get("resort name")
        activity_col = header_map.get("activity")
        bookable_col = header_map.get("bookable hours")
        month_col = header_map.get("month")
        duration_col = header_map.get("activity duration")

        if not (month_col and activity_col and bookable_col):
            # missing required columns -> skip
            continue

        date_start_col = month_col + 1
        max_check_col = min(ws_src.max_column, date_start_col + 31 - 1)
        out_row = 2

        # detect multi-resort activities (to attach resort to event name when needed)
        activity_resorts = {}
        for r in range(2, ws_src.max_row + 1):
            act = safe_str(ws_src.cell(r, activity_col).value)
            res = safe_str(ws_src.cell(r, resort_col).value) if resort_col else ""
            if act:
                activity_resorts.setdefault(act, set()).add(res)

        seen_events = set()

        for r in range(2, ws_src.max_row + 1):
            activity = safe_str(ws_src.cell(r, activity_col).value)
            if not activity:
                continue
            resort = safe_str(ws_src.cell(r, resort_col).value) if resort_col else ""
            duration = safe_str(ws_src.cell(r, duration_col).value) if duration_col else ""
            month_val = ws_src.cell(r, month_col).value

            # skip multi-day Galaxea
            if sheet_name.upper() == "GALAXEA" and "day" in duration.lower():
                continue

            # find first non-empty day (original logic)
            first_day = None
            for col in range(date_start_col, max_check_col + 1):
                c = ws_src.cell(r, col)
                try:
                    if c.value is not None and float(c.value) > 0:
                        first_day = int(ws_src.cell(1, col).value)
                        break
                except Exception:
                    continue
            if not first_day:
                continue

            month_num = None
            try:
                month_num = int(month_val)
            except Exception:
                # try parsing textual month
                month_num = parse_month_to_num(month_val)
            if not month_num:
                continue
            date_str = f"{first_day:02d}/{month_num:02d}/{YEAR_FOR_OUTPUT}"

            resorts_for_activity = activity_resorts.get(activity, set())
            if len(resorts_for_activity) > 1 and resort:
                event_name = f"{activity} - {resort}"
            else:
                event_name = activity

            instrs = instructors_map.get(sheet_name, {}).get(activity, [])

            if event_name not in event_color_cache:
                event_color_cache[event_name] = get_light_fill()
            fill = event_color_cache[event_name]

            # determine bookable cell coordinate (e.g., "E5")
            bookable_coord = f"{get_column_letter(bookable_col)}{r}"

            # get formula for this specific cell (from XML-parsed dv_map)
            sheet_dvs = dv_map.get(sheet_name, {})
            formula = sheet_dvs.get(bookable_coord)

            time_slots = []
            if formula:
                # resolve using openpyxl workbook
                time_slots = resolve_formula_to_values(wb_src, formula)
            # else: no dataValidation found for this particular cell -> no slots

            # populate rows for each timeslot
            for slot in time_slots:
                slot = slot.strip()
                if not slot:
                    continue
                if "-" in slot:
                    parts = [p.strip() for p in slot.split("-", 1)]
                    start_time = parts[0]
                    end_time = parts[1] if len(parts) > 1 else ""
                else:
                    start_time = slot
                    end_time = ""

                key = (event_name, resort, activity, date_str, start_time, end_time)
                if key in seen_events:
                    continue
                seen_events.add(key)

                # main event row
                row_vals = [event_name, activity, activity, date_str, start_time, end_time]
                for col_idx, val in enumerate(row_vals, start=1):
                    ws_out.cell(row=out_row, column=col_idx, value=val).fill = fill
                out_row += 1

                # instructor rows
                for instr in instrs:
                    instr_vals = [event_name, instr, activity, date_str, start_time, end_time]
                    for col_idx, val in enumerate(instr_vals, start=1):
                        ws_out.cell(row=out_row, column=col_idx, value=val).fill = fill
                    out_row += 1

    wb_out.save(output_file)
    print(f"✅ Output saved to {output_file}")
